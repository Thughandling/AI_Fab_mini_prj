#!/usr/bin/env python3
"""AI_DigitalFactory_개발정의.xlsx — 웨이퍼 Agent 기준으로 재작성."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "AI_DigitalFactory_개발정의.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, headers, rows, col_widths=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    for row in rows:
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(ws.max_row, c).alignment = WRAP
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 개발정의 ──
    ws = wb.create_sheet("개발정의")
    write_sheet(
        ws,
        ["셀/모듈", "함수/요소", "목적", "인자/설명"],
        [
            ["Cell 0", "-", "노트북 개요 — Digital Factory 웨이퍼 이상탐지 Agent", "-"],
            ["Cell 1", "pip install", "의존성 설치 (Colab)", "transformers>=4.51, torch, accelerate, langgraph, langchain, pandas, statsmodels ..."],
            ["Cell 2", "GPU 확인", "Colab T4 GPU 사용 여부 확인", "torch.cuda.is_available()"],
            ["Cell 3", "Drive mount", "Colab Drive 마운트 (선택)", "google.colab.drive"],
            ["Cell 4", "LLM_PROVIDER", "LLM 백엔드 선택", "mock | qwen | anthropic"],
            ["Cell 4", "HF_MODEL_ID", "HuggingFace 모델 ID", "Qwen/Qwen3-0.6B (기본)"],
            ["Cell 4", "N_STEPS / Z_SCORE", "시뮬레이션·탐지 파라미터", "500 steps, z=1.96 (95% CI)"],
            ["Cell 5", "load_qwen_model()", "Qwen3-0.6B 로드 (GPU/CPU 자동)", "model_id, trust_remote_code=True"],
            ["Cell 5", "_generate_qwen()", "chat template 기반 텍스트 생성", "system, user, max_new_tokens=384"],
            ["Cell 5", "invoke_structured()", "JSON 구조화 LLM 호출", "system, user, Pydantic schema → dict"],
            ["Cell 6", "generate_sensor_data()", "ARMA(1,1) + 이상 주입 시뮬레이터", "n_steps=500, seed=42"],
            ["Cell 6", "_inject_spike/drift/loss", "이상 패턴 주입", "SPIKE(장비), DRIFT(공정), LOSS(센서)"],
            ["Cell 7", "run_detection()", "Tool1 — ARMA 이상탐지", "history DataFrame, current row → is_anomaly"],
            ["Cell 7", "detect_anomaly @tool", "LangChain Tool 래퍼", "sensor_history_json, current_point_json"],
            ["Cell 7", "run_classification()", "Tool2 — 원인 분류 (LLM)", "recent 20 points → SPIKE|DRIFT|LOSS"],
            ["Cell 7", "classify_anomaly_cause @tool", "LangChain Tool 래퍼", "recent_data_json"],
            ["Cell 7", "run_action_decision()", "Tool3 — 대응 Level 1~3 결정", "classification dict → action_level"],
            ["Cell 7", "decide_action @tool", "LangChain Tool 래퍼", "classification_json"],
            ["Cell 7", "run_report()", "Tool4 — Markdown 리포트 생성", "anomaly, classification, action → .md"],
            ["Cell 7", "generate_report @tool", "LangChain Tool 래퍼", "anomaly/classification/action JSON"],
            ["Cell 8", "WaferState", "LangGraph 공유 State (TypedDict)", "sensor_data, current_idx, anomaly_result, ..."],
            ["Cell 8", "detect_node", "이상탐지 노드", "run_detection() 호출"],
            ["Cell 8", "classify_node", "원인 분류 노드", "run_classification() + Qwen"],
            ["Cell 8", "action_node", "대응 결정 노드", "run_action_decision() + Qwen"],
            ["Cell 8", "report_gen", "리포트 생성 노드", "run_report() → reports/"],
            ["Cell 8", "route_after_detect", "조건부 엣지", "is_anomaly → classify | end"],
            ["Cell 9", "모니터링 루프", "500 step 순차 실행 + debounce", "warmup=30, cooldown=25"],
            ["Cell 10", "Qwen 단건 테스트", "분류·대응 LLM 단건 확인", "run_classification, run_action_decision"],
            ["Cell 11", "시각화", "matplotlib 차트 + GT 오버레이", "events, sensor_df"],
        ],
        col_widths=[14, 28, 36, 48],
    )

    ws2 = wb.create_sheet("변경이력")
    write_sheet(
        ws2,
        ["일자", "항목", "변경 내용"],
        [
            ["2026-07-05", "프로젝트 전환", "AI SecOps → Digital Factory 웨이퍼 이상탐지 & 자동대응 Agent"],
            ["2026-07-05", "LLM", "HuggingFace Qwen3-0.6B (Colab GPU) + mock/anthropic 지원"],
            ["2026-07-05", "LangGraph", "detect → classify → action → report_gen 파이프라인"],
            ["2026-07-05", "시뮬레이터", "ARMA(1,1) + SPIKE/DRIFT/LOSS 이상 주입"],
            ["2026-07-05", "Colab", "GPU 확인, transformers 4.51+, Qwen 로드 셀 추가"],
            ["2026-07-05", "구조 단순화", "Colab 전용 — wafer_agent/ .py 패키지 제거, 노트북 단일화"],
        ],
        col_widths=[14, 22, 60],
    )

    ws3 = wb.create_sheet("LLM_Provider_비교")
    write_sheet(
        ws3,
        ["Provider", "구현", "환경변수", "Colab", "용도"],
        [
            ["mock", "규칙 기반 _mock_classify/_mock_action", "LLM_PROVIDER=mock", "Secret 불필요", "파이프라인·구조 검증"],
            ["qwen", "HuggingFace Qwen3-0.6B 로컬 추론", "LLM_PROVIDER=qwen, HF_MODEL_ID", "T4 GPU 권장", "Colab 실제 LLM 분류/대응"],
            ["anthropic", "langchain_anthropic ChatAnthropic", "ANTHROPIC_API_KEY", "ANTHROPIC_API_KEY Secret", "Claude API 기반 분석"],
        ],
        col_widths=[14, 32, 36, 22, 28],
    )

    ws4 = wb.create_sheet("Mock모드_설명")
    write_sheet(
        ws4,
        ["항목", "설명"],
        [
            ["Mock이란?", "Qwen/Claude API 없이 규칙 기반으로 SPIKE/DRIFT/LOSS 분류 및 Level 1~3 대응"],
            ["선택", "Cell 4: LLM_PROVIDER = \"mock\""],
            ["분류 로직", "결측→LOSS, 급변→SPIKE, 기울기→DRIFT 휴리스틱"],
            ["대응 로직", "LOSS+고신뢰→Level3, SPIKE/DRIFT→Level2, 그 외→Level1"],
            ["한계", "실제 LLM 추론 없음 — 구조 검증·데모용"],
            ["실전", "Colab: LLM_PROVIDER=\"qwen\" + GPU"],
        ],
        col_widths=[18, 70],
    )

    ws5 = wb.create_sheet("파이프라인_플로우")
    write_sheet(
        ws5,
        ["순서", "노드", "입력", "출력", "LLM"],
        [
            ["0", "simulator", "-", "sensor_data (500 steps)", "X"],
            ["1", "detect", "history + current point", "anomaly_result (is_anomaly, CI)", "X (ARMA)"],
            ["-", "조건부", "is_anomaly=False", "END (모니터링 계속)", "-"],
            ["2", "classify", "최근 20 시점", "classification_result (SPIKE/DRIFT/LOSS)", "O (Qwen/mock)"],
            ["3", "action", "classification", "action_result (Level 1~3)", "O (Qwen/mock)"],
            ["4", "report_gen", "anomaly+classification+action", "anomaly_report_{ts}.md", "X"],
        ],
        col_widths=[8, 14, 28, 36, 16],
    )

    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
